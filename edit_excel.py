import os
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import pandas as pd

def remove_colon(value):
    if isinstance(value, str) and value.startswith(':'):
        return value[1:]  # Remove the first character (colon)
    return value

def preprocess(df):
  df = df.applymap(remove_colon)
  # Extract the first 3 words and put them in a new column
  df['Key'] = df['Sec_no_vill'].str.split().str[:1].str.join(' ')

  # Move the 'unique_key' column to the first position
  df.insert(0, 'Key', df.pop('Key'))
  return df

def get_headers(dashboard_sheet):
  # Define column headers
  column_headers = ['Location', 'Village', 'Male', 'Female', 'Total (M+F)', 'Age(18-30)', 'Age(31-45)', 'Age(46-60)', 'Age(60+)']

  start_column_index = 4
  # Add column headers to the 5th row, starting from the 4th column
  for header in column_headers:
      cell = dashboard_sheet.cell(row=5, column=start_column_index)
      cell.value = header
      start_column_index += 1  # Increment the column index for the next header

def get_unique_locations(sheet1, dashboard_sheet):
  # Extract unique values from the 'Key' column in 'Sheet1'
  unique_values = set()
  for row in sheet1.iter_rows(min_row=2, min_col=1, values_only=True):
      key_value = row[0]  # Assuming 'Key' column is the first column (column 1)
      if key_value:
          unique_values.add(key_value)

  # Sort the unique values based on custom sorting
  sorted_unique_values = sorted(unique_values, key=lambda x: (int(x.split('-')[0]) if x.split('-')[0].isdigit() else float('inf'), x))

  # Write the sorted values to the 'Location' column in 'Dashboard'
  for i, value in enumerate(sorted_unique_values, start=6):  # Start from the 6th row
      cell = dashboard_sheet.cell(row=i, column=4)  # Assuming 'Location' column is the 4th column (column D)
      cell.value = value

def update_male_female(dashboard_sheet, sheet1):
    # Get the unique locations from the Dashboard sheet
    unique_locations = [cell.value for cell in dashboard_sheet['D'] if cell.value and cell.row > 5]

    # Iterate through unique locations and count 'पुरुष' and 'पुरुध' (male) and 'महिला' (female)
    for location in unique_locations:
        male_count = 0
        female_count = 0
        for cell in sheet1['A']:
            if cell.value == location:
                # Get the corresponding 'Sex' value in the same row
                sex_value = sheet1.cell(row=cell.row, column=7).value  # Assuming 'Sex' is in column G (column index 7)
                if sex_value in ['पुरुष', 'पुरुध']:
                    male_count += 1
                elif sex_value == 'महिला':
                    female_count += 1

        # After processing all rows with the same location, write the counts to the Dashboard sheet
        for cell in dashboard_sheet['D']:
            if cell.value == location:
                dashboard_sheet.cell(row=cell.row, column=6).value = male_count  # Write male count to 'Male' column (column 6)
                dashboard_sheet.cell(row=cell.row, column=7).value = female_count  # Write female count to 'Female' column (column 7)

def update_total_column(dashboard_sheet):
    # Initialize the starting row number
    current_row = 6

    # Iterate through rows starting from row 6 (assuming the data starts from row 6)
    for row in dashboard_sheet.iter_rows(min_row=current_row, values_only=True):
        male_value = row[5]  # Assuming 'Male' column is the 6th column (column F)
        female_value = row[6]  # Assuming 'Female' column is the 7th column (column G)

        # Initialize the total_value variable
        total_value = 0

        # Check if male_value and female_value are not None before performing addition
        if male_value is not None and female_value is not None:
            # Calculate the sum of male and female values
            total_value = male_value + female_value

            # Update the 'Total (M+F)' column (assuming it's the 8th column, column H)
            dashboard_sheet.cell(row=current_row, column=8, value=total_value)

        # Increment the row number for the next iteration
        current_row += 1

def print_sheet(dashboard_sheet):
  for row in dashboard_sheet.iter_rows(values_only=True):
    print(row)


def update_age_columns(dashboard_sheet, df):
    unique_locations = df['Key'].unique().tolist()
    age_counts = {}  # Dictionary to store age_count for each location

    # Calculate age counts for each location
    for location in unique_locations:
        age_count_18_to_30 = len(df[(df['Age'] > 17) & (df['Age'] < 31) & (df['Key'] == location)])
        age_count_31_to_45 = len(df[(df['Age'] > 30) & (df['Age'] < 46) & (df['Key'] == location)])
        age_count_46_to_60 = len(df[(df['Age'] > 45) & (df['Age'] < 61) & (df['Key'] == location)])
        age_count_60_plus = len(df[(df['Age'] > 60) & (df['Key'] == location)])

        age_counts[location] = {
            '18-30': age_count_18_to_30,
            '31-45': age_count_31_to_45,
            '46-60': age_count_46_to_60,
            '60+': age_count_60_plus
        }

    # Update the corresponding cells in the Excel sheet
    current_row = 6  # Start from the 6th row as per your example
    for row in dashboard_sheet.iter_rows(min_row=current_row, values_only=True):
        location = row[3]  # Assuming 'Location' is in the 4th column (0-based indexing)
        print("Processing location:", location)

        if location in age_counts:
            age_data = age_counts[location]
            print("Age data for location:", age_data)
            # Update the age group columns in the 'dashboard_sheet'
            dashboard_sheet.cell(row=current_row, column=9, value=age_data['18-30'])
            dashboard_sheet.cell(row=current_row, column=10, value=age_data['31-45'])
            dashboard_sheet.cell(row=current_row, column=11, value=age_data['46-60'])
            dashboard_sheet.cell(row=current_row, column=12, value=age_data['60+'])

        current_row += 1   


src_dir = 'D:/Amber_AC_Final_Revision/Amber_Final_Voter_List_Excel'
temp_dir = 'D:/Amber_AC_Final_Revision/temp'
dst_dir = 'D:/Amber_AC_Final_Revision/Amber_Final_Voter_List_Excel_Modified'

# Get excel file names from the 'src_dir' and save it in a list
excel_files_src = [f for f in os.listdir(src_dir) if f.lower().endswith('.xlsx')]

# Loop and Get excel files from the 'src_dir', process and store in 'temp_dir'
for excel_file in excel_files_src:
    print("File No:{}".format(excel_file))
    excel_path = os.path.join(src_dir, excel_file)
    df = pd.read_excel(excel_path, sheet_name='Sheet1')
    df1 = preprocess(df)
    excel_path = os.path.join(temp_dir, excel_file)
    df1.to_excel(excel_path, index=False)
    print('File {} saved to temporary folder!'.format(excel_file))


# Get excel file names from the 'temp_dir' and save it in a list
excel_files_temp = [f for f in os.listdir(temp_dir) if f.lower().endswith('.xlsx')]

# Loop and get excel file from 'temp_dir', process and store in 'dest_dir'
for excel_file in excel_files_temp:
    print("File No:{}".format(excel_file))
    excel_path = os.path.join(temp_dir, excel_file)
    df = pd.read_excel(excel_path, sheet_name='Sheet1')
    # Load the Excel workbook
    wb = load_workbook(excel_path)
    # Create a new sheet named "Dashboard"
    dashboard_sheet = wb.create_sheet(title='Dashboard')
    # Select the Sheet1
    sheet1 = wb['Sheet1']
    # Select the Dashboard sheet
    dashboard_sheet = wb['Dashboard']

    get_headers(dashboard_sheet)

    get_unique_locations(sheet1, dashboard_sheet)

    update_male_female(dashboard_sheet, sheet1)

    update_total_column(dashboard_sheet)

    print("Entering Age columns!\n\n")
    update_age_columns(dashboard_sheet, df)
    print_sheet(dashboard_sheet)

    excel_path = os.path.join(dst_dir, excel_file)
    wb.save(excel_path)
    print("File {} saved to destination folder!".format(excel_file))

print("Entire Operation Completed!")